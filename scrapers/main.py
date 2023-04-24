import os
import pandas as pd
from openpyxl import load_workbook

def save_into_excel_db(item_dict, store_name, store_type, save_type='append'):
    parent_dir = os.path.abspath('..')
    file_name = f'{parent_dir}/Efficient Eats/nutrition_dict.xlsx'
    df = pd.DataFrame.from_dict(item_dict, orient='index')
    df['Item Name'] = df.index
    cols = df.columns.tolist()
    cols = [cols[-1]] + cols[:-1]
    df = df[cols]
    df['Store Name'] = store_name
    df['Store Type'] = store_type
        
    if os.path.isfile(file_name): 
        if save_type == 'append':
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
                writer.book = load_workbook(file_name)
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                df.to_excel(writer, sheet_name=store_name, startrow=writer.sheets[store_name].max_row, index=False, header=False)
        elif save_type == 'overwrite':
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
                if store_name in writer.book.sheetnames:
                    writer.book.remove_sheet(writer.book[store_name])
                df.to_excel(writer, sheet_name=store_name, index=False)
    
    else:
        df.to_excel(file_name, sheet_name=store_name, index=False)